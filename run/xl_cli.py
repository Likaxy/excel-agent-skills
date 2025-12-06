#!/usr/bin/env python3
"""CLI to run Excel ops (same logic as server, no MCP). Usage: python xl_cli.py <op> [args...]. Paths must be absolute."""
import inspect
import json
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

def _resolve(path: str) -> str:
    return path if os.path.isabs(path) else os.path.abspath(path)

def _coerce(val, param):
    if param.annotation != inspect.Parameter.empty:
        if param.annotation == bool:
            return str(val).lower() in ("1", "true", "yes")
        if param.annotation == int:
            return int(val)
        if "List" in str(param.annotation) or "Dict" in str(param.annotation):
            return json.loads(val) if isinstance(val, str) else val
    return val

def main():
    from xl_agent import workbook, data
    from xl_agent.sheet import (
        copy_sheet, delete_sheet, rename_sheet, merge_range, unmerge_range,
        get_merged_ranges, copy_range_operation, delete_range_operation,
        insert_row, insert_cols, delete_rows, delete_cols,
    )
    from xl_agent.validation import validate_range_in_sheet_operation
    from xl_agent.chart import create_chart_in_sheet as create_chart_impl
    from xl_agent.pivot import create_pivot_table as create_pivot_table_impl
    from xl_agent.tables import create_excel_table as create_table_impl
    from xl_agent.calculations import apply_formula as apply_formula_impl
    from xl_agent.validation import validate_formula_in_cell_operation as validate_formula_impl
    from xl_agent.formatting import format_range as format_range_func
    from xl_agent.data import read_excel_range_with_metadata

    OPS = [
        "apply_formula", "validate_formula_syntax", "format_range", "read_data_from_excel", "write_data_to_excel",
        "create_workbook", "create_worksheet", "create_chart", "create_pivot_table", "create_table",
        "copy_worksheet", "delete_worksheet", "rename_worksheet", "get_workbook_metadata",
        "merge_cells", "unmerge_cells", "get_merged_cells", "copy_range", "delete_range",
        "validate_excel_range", "get_data_validation_info", "insert_rows", "insert_columns",
        "delete_sheet_rows", "delete_sheet_columns",
    ]
    if len(sys.argv) < 2 or sys.argv[1] not in OPS:
        print("Usage: xl_cli.py <op> [args...]", file=sys.stderr)
        sys.exit(1)
    op = sys.argv[1]
    args_in = sys.argv[2:]

    handlers = {
        "create_workbook": lambda: workbook.create_workbook(_resolve(args_in[0]))["message"],
        "create_worksheet": lambda: workbook.create_sheet(_resolve(args_in[0]), args_in[1])["message"],
        "get_workbook_metadata": lambda: str(workbook.get_workbook_info(_resolve(args_in[0]), include_ranges=len(args_in) > 1 and args_in[1].lower() == "true")),
        "write_data_to_excel": lambda: data.write_data(_resolve(args_in[0]), args_in[1], json.loads(args_in[2]) if len(args_in) > 2 else [], args_in[3] if len(args_in) > 3 else "A1")["message"],
        "read_data_from_excel": lambda: json.dumps(data.read_excel_range_with_metadata(_resolve(args_in[0]), args_in[1], args_in[2] if len(args_in) > 2 else "A1", args_in[3] if len(args_in) > 3 else None), indent=2, default=str),
        "apply_formula": lambda: apply_formula_impl(_resolve(args_in[0]), args_in[1], args_in[2], args_in[3])["message"],
        "validate_formula_syntax": lambda: validate_formula_impl(_resolve(args_in[0]), args_in[1], args_in[2], args_in[3])["message"],
        "format_range": lambda: (format_range_func(filepath=_resolve(args_in[0]), sheet_name=args_in[1], start_cell=args_in[2], end_cell=args_in[3] if len(args_in) > 3 else None) or {}, "Range formatted successfully")[1],
        "merge_cells": lambda: merge_range(_resolve(args_in[0]), args_in[1], args_in[2], args_in[3])["message"],
        "unmerge_cells": lambda: unmerge_range(_resolve(args_in[0]), args_in[1], args_in[2], args_in[3])["message"],
        "get_merged_cells": lambda: str(get_merged_ranges(_resolve(args_in[0]), args_in[1])),
        "copy_worksheet": lambda: copy_sheet(_resolve(args_in[0]), args_in[1], args_in[2])["message"],
        "delete_worksheet": lambda: delete_sheet(_resolve(args_in[0]), args_in[1])["message"],
        "rename_worksheet": lambda: rename_sheet(_resolve(args_in[0]), args_in[1], args_in[2])["message"],
        "copy_range": lambda: copy_range_operation(_resolve(args_in[0]), args_in[1], args_in[2], args_in[3], args_in[4], args_in[5] if len(args_in) > 5 else args_in[1])["message"],
        "delete_range": lambda: delete_range_operation(_resolve(args_in[0]), args_in[1], args_in[2], args_in[3], args_in[4] if len(args_in) > 4 else "up")["message"],
        "validate_excel_range": lambda: validate_range_in_sheet_operation(_resolve(args_in[0]), args_in[1], args_in[2], args_in[3] if len(args_in) > 3 else None)["message"],
        "insert_rows": lambda: insert_row(_resolve(args_in[0]), args_in[1], int(args_in[2]), int(args_in[3]) if len(args_in) > 3 else 1)["message"],
        "insert_columns": lambda: insert_cols(_resolve(args_in[0]), args_in[1], int(args_in[2]), int(args_in[3]) if len(args_in) > 3 else 1)["message"],
        "delete_sheet_rows": lambda: delete_rows(_resolve(args_in[0]), args_in[1], int(args_in[2]), int(args_in[3]) if len(args_in) > 3 else 1)["message"],
        "delete_sheet_columns": lambda: delete_cols(_resolve(args_in[0]), args_in[1], int(args_in[2]), int(args_in[3]) if len(args_in) > 3 else 1)["message"],
        "create_chart": lambda: create_chart_impl(filepath=_resolve(args_in[0]), sheet_name=args_in[1], data_range=args_in[2], chart_type=args_in[3], target_cell=args_in[4], title=args_in[5] if len(args_in) > 5 else "", x_axis=args_in[6] if len(args_in) > 6 else "", y_axis=args_in[7] if len(args_in) > 7 else "")["message"],
        "create_pivot_table": lambda: create_pivot_table_impl(_resolve(args_in[0]), args_in[1], args_in[2], json.loads(args_in[3]) if len(args_in) > 3 else [], json.loads(args_in[4]) if len(args_in) > 4 else [], json.loads(args_in[5]) if len(args_in) > 5 else None, args_in[6] if len(args_in) > 6 else "mean")["message"],
        "create_table": lambda: create_table_impl(_resolve(args_in[0]), args_in[1], args_in[2], args_in[3] if len(args_in) > 3 else None, args_in[4] if len(args_in) > 4 else "TableStyleMedium9")["message"],
    }
    if op == "get_data_validation_info":
        from openpyxl import load_workbook
        from xl_agent.cell_validation import get_all_validation_ranges
        wb = load_workbook(_resolve(args_in[0]), read_only=False)
        ws = wb[args_in[1]]
        validations = get_all_validation_ranges(ws)
        wb.close()
        out = json.dumps({"sheet_name": args_in[1], "validation_rules": validations}, indent=2, default=str) if validations else "No data validation rules found in this worksheet"
    elif op not in handlers:
        print(f"Unknown op: {op}", file=sys.stderr)
        sys.exit(1)
    else:
        out = handlers[op]()
    print(out)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
